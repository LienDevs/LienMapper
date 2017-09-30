# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.shortcuts import render

# Create your views here.


from django.http import HttpResponse

from .models import Property

import openpyxl as xl
from AddressToLatLon import findLatLng
XLSNAME = 'liens.xlsx'

def index(request):
	updateDatabase(XLSTODB('C:\\Users\\jraed\\lienin\\lienMapper\\liens.xlsx'))
	return HttpResponse("Success")

def XLSTODB(XLSNAME):
    

	wb = xl.load_workbook(filename = XLSNAME)

	sheet = wb['Sheet1']

	fileedition = sheet.cell(row = 1, column = 2).value
	STATE = sheet.cell(row = 2, column = 2).value
	CITY = sheet.cell(row = 3, column = 2).value
	SALEDATE = sheet.cell(row = 4, column = 2).value
	row = 6 
	ID = sheet.cell(row = row, column = 1)
	OWNER = sheet.cell(row = row, column = 2)
	ADDRESS = sheet.cell(row = row, column = 3)
	BID = sheet.cell(row = row, column = 4)
	rows = []
	##puts into list of tuples per row 
	for i in range(408):
		##function grabs lat and long from current address

		LAT = findLatLng(ADDRESS.value + " " + CITY + " " + STATE)[0]
		LNG = findLatLng(ADDRESS.value + " " + CITY + " " + STATE)[1]

		##appends all data into temp list of tuples
		rows.append((ID.value,OWNER.value,ADDRESS.value,STATE,CITY,BID.value,SALEDATE,LAT,LNG))
		print str(row) + " done"
		row += 1
		ID = sheet.cell(row = row, column = 1)
		OWNER = sheet.cell(row = row, column = 2)
		ADDRESS = sheet.cell(row = row, column = 3)
		BID = sheet.cell(row = row, column = 4)
	wb.close()
	return rows

	
def updateDatabase(rows):

	for i in rows:
		id = str(i[0])
		owner = str(i[1])
		address = str(i[2])
		state = str(i[3])
		city = str(i[4])
		bid = str(i[5])
		saledate = str(i[6])
		lat = str(i[7])
		lng = str(i[8])
	

		new_property = Property.objects.create(prop_id = id, owner=owner, address=address, state=state, city=city, bid=bid, date=saledate, lat=lat, lng=lng)

	new_property.save()
		
		


	
	
	
			